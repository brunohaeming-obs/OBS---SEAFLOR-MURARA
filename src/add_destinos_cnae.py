"""
Reconstrói a aba "14_Dest_CNAE" no SEAFLOR_2026_JanMai_3anos_v2.xlsx.

Estrutura: País (linha total) → breakdown por CNAE abaixo.
Fonte:
  - 5_Top_Destinos → totais do Complexo por país (top-15 por 2026)
  - 9_Dest_Setor   → breakdown por CNAE/país
"""
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX     = Path(r"c:\Users\Janine\OneDrive\Área de Trabalho\OBS---SEAFLOR-MURARA\data\processed\SEAFLOR_2026_JanMai_3anos_v2.xlsx")
XLSX_SRC = Path(r"C:\Users\Janine\AppData\Local\Temp\seaflor_copia.xlsx")  # leitura (cópia temp)
XLSX_OUT = Path(r"C:\Users\Janine\AppData\Local\Temp\seaflor_14_novo.xlsx")  # saída temporária

# Paleta
AZUL_ESC = "1F4E79"
AZUL_MED = "2E75B6"
AZUL_CL  = "DEEAF1"
VERDE    = "375623"
VERMELHO = "C00000"
BRANCO   = "FFFFFF"
CINZA_CL = "F2F2F2"
AMARELO  = "FFD966"

SETORES_ORDEM = [
    ("Madeira (CNAE 16)",          "CNAE 16 — Madeira"),
    ("Moveis (CNAE 31)",           "CNAE 31 — Móveis"),
    ("Papel e Celulose (CNAE 17)", "CNAE 17 — Papel e Celulose"),
    ("Base Florestal (CNAE 2)",    "CNAE 2  — Base Florestal"),
]

TITULO = ("14. Exportações por País e CNAE — Complexo Florestal SC"
          " | Jan-Mai 2024 | 2025 | 2026  (top-15 destinos por valor 2026)")

COLUNAS = [
    ("Destino / Segmento",       32),
    ("EXP Jan-Mai 2024 (US$)",   18),
    ("EXP Jan-Mai 2025 (US$)",   18),
    ("Var 24/25 (%)",            12),
    ("EXP Jan-Mai 2026 (US$)",   18),
    ("Var 25/26 (%)",            12),
]


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _border() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _font(bold=False, color=None, size=9, italic=False) -> Font:
    return Font(bold=bold, size=size, color=(color or "000000"), italic=italic)

def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def var_pct(new, old):
    try:
        if old and old != 0:
            return round((new - old) / abs(old) * 100, 1)
    except Exception:
        pass
    return None


def load_data():
    print("Lendo dados do Excel...")
    # Usa cópia temporária se o original estiver bloqueado pelo Excel
    src = XLSX_SRC if XLSX_SRC.exists() else XLSX
    print(f"  Fonte: {src}")
    f = str(src)

    # ── 5_Top_Destinos → totais do Complexo ──────────────────────────────────
    df5 = pd.read_excel(f, sheet_name="5_Top_Destinos", header=1)
    df5 = df5[df5["Pais"].notna()].copy()
    top_paises = df5.nsmallest(15, "Rank_2026")["Pais"].tolist()
    complex_data = (df5.set_index("Pais")
                      [["EXP_USD_2024", "EXP_USD_2025", "EXP_USD_2026",
                        "Var_EXP_24_25_Pct", "Var_EXP_25_26_Pct"]])

    # ── 9_Dest_Setor → breakdown por CNAE/país ───────────────────────────────
    df9 = pd.read_excel(f, sheet_name="9_Dest_Setor", header=1)
    df9 = df9[df9["Setor"].notna() & df9["Pais"].notna()].copy()

    setor_lookup = {}
    for setor, _ in SETORES_ORDEM:
        sub = df9[df9["Setor"] == setor].copy()
        setor_lookup[setor] = sub.set_index("Pais")[["EXP_2024", "EXP_2025", "EXP_2026"]]

    print(f"  {len(top_paises)} destinos do Complexo lidos")
    return top_paises, complex_data, setor_lookup


def build_rows(top_paises, complex_data, setor_lookup):
    rows = []  # list of dicts: {label, tipo, e24, e25, v2425, e26, v2526, rank}
    for rank, pais in enumerate(top_paises, 1):
        # linha do país (total Complexo)
        row_c = complex_data.loc[pais] if pais in complex_data.index else {}
        rows.append({
            "rank":  rank,
            "label": pais,
            "tipo":  "pais",
            "e24":   int(row_c.get("EXP_USD_2024", 0) or 0),
            "e25":   int(row_c.get("EXP_USD_2025", 0) or 0),
            "v2425": row_c.get("Var_EXP_24_25_Pct", None),
            "e26":   int(row_c.get("EXP_USD_2026", 0) or 0),
            "v2526": row_c.get("Var_EXP_25_26_Pct", None),
        })
        # linhas dos CNAEs
        for setor_key, setor_label in SETORES_ORDEM:
            lookup = setor_lookup.get(setor_key, pd.DataFrame())
            if pais in lookup.index:
                r = lookup.loc[pais]
                e24 = int(r["EXP_2024"] or 0)
                e25 = int(r["EXP_2025"] or 0)
                e26 = int(r["EXP_2026"] or 0)
                v2425 = var_pct(e25, e24)
                v2526 = var_pct(e26, e25)
            else:
                e24 = e25 = e26 = 0
                v2425 = v2526 = None
            rows.append({
                "rank":  rank,
                "label": f"    {setor_label}",
                "tipo":  "cnae",
                "e24":   e24,
                "e25":   e25,
                "v2425": v2425,
                "e26":   e26,
                "v2526": v2526,
            })
    return rows


def write_sheet(ws, rows, wb):
    n_cols = len(COLUNAS)

    # Linha 1: título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(1, 1, TITULO)
    c.font      = _font(bold=True, size=11, color=BRANCO)
    c.fill      = _fill(AZUL_MED)
    c.alignment = _align("left", "center")
    ws.row_dimensions[1].height = 22

    # Linha 2: cabeçalhos
    ws.row_dimensions[2].height = 36
    for ci, (label, width) in enumerate(COLUNAS, 1):
        c = ws.cell(2, ci, label)
        c.font      = _font(bold=True, size=9, color=BRANCO)
        c.fill      = _fill(AZUL_ESC)
        c.alignment = _align("center", "center", wrap=True)
        c.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = width

    cur_row = 3
    for r in rows:
        is_pais = r["tipo"] == "pais"
        ws.row_dimensions[cur_row].height = 16 if is_pais else 13

        # Determinar cores por rank (alternado a cada país)
        if is_pais:
            bg_label  = AZUL_MED
            bg_num    = AZUL_CL
            ft_label  = _font(bold=True, size=10, color=BRANCO)
            ft_num    = _font(bold=True, size=9)
        else:
            even = r["rank"] % 2 == 0
            bg_label = AZUL_CL if even else BRANCO
            bg_num   = AZUL_CL if even else BRANCO
            ft_label = _font(italic=True, size=9)
            ft_num   = _font(size=9)

        # Col 1: nome
        c1 = ws.cell(cur_row, 1, r["label"])
        c1.font      = ft_label
        c1.fill      = _fill(bg_label)
        c1.alignment = _align("left", "center")
        c1.border    = _border()

        # Colunas numéricas
        vals = [
            (2, r["e24"],   "#,##0",   None),
            (3, r["e25"],   "#,##0",   None),
            (4, r["v2425"], '0.0"%"',  True),
            (5, r["e26"],   "#,##0",   None),
            (6, r["v2526"], '0.0"%"',  True),
        ]
        for ci, val, nfmt, is_var in vals:
            c = ws.cell(cur_row, ci)
            c.border    = _border()
            c.fill      = _fill(bg_num)
            c.alignment = _align("right", "center")

            if val is None or (isinstance(val, float) and pd.isna(val)):
                c.value = ""
                c.font  = ft_num
            elif is_var:
                c.value         = val
                c.number_format = nfmt
                if isinstance(val, (int, float)):
                    cor = VERDE if val > 0 else (VERMELHO if val < 0 else "000000")
                else:
                    cor = "000000"
                c.font = _font(bold=is_pais, size=(10 if is_pais else 9), color=cor)
            else:
                c.value         = val if val != 0 else ""
                c.number_format = nfmt
                c.font          = ft_num

        cur_row += 1

    ws.freeze_panes = "B3"

    # Nota de rodapé
    ws.row_dimensions[cur_row + 1].height = 30
    ws.merge_cells(start_row=cur_row + 1, start_column=1,
                   end_row=cur_row + 1, end_column=n_cols)
    nota = (
        "Nota: Totais por país = Complexo Florestal (CNAE 2+16+17+31, aba 5_Top_Destinos). "
        "CNAEs em branco = país fora do Top-15 daquele segmento (9_Dest_Setor). "
        "Valores em US$ FOB Jan-Mai."
    )
    cn = ws.cell(cur_row + 1, 1, nota)
    cn.font      = _font(italic=True, size=8, color="595959")
    cn.alignment = _align("left", "center", wrap=True)


def main():
    top_paises, complex_data, setor_lookup = load_data()
    rows = build_rows(top_paises, complex_data, setor_lookup)

    print(f"Gerando {len(rows)} linhas ({len(top_paises)} países × 5 linhas)...")

    # Abre a cópia para preservar todas as abas existentes
    src = XLSX_SRC if XLSX_SRC.exists() else XLSX
    wb = load_workbook(str(src))
    if "14_Dest_CNAE" in wb.sheetnames:
        del wb["14_Dest_CNAE"]
        print("  Aba anterior removida.")

    ws = wb.create_sheet("14_Dest_CNAE")
    write_sheet(ws, rows, wb)

    # Mover para última posição (após aba 13)
    if "13_Dest_Detalhe" in wb.sheetnames:
        idx_13 = wb.sheetnames.index("13_Dest_Detalhe")
        wb.move_sheet("14_Dest_CNAE", offset=idx_13 + 1 - (len(wb.sheetnames) - 1))

    # Salva: tenta direto; se bloqueado, salva em temp
    try:
        wb.save(str(XLSX))
        print(f"\nArquivo salvo: {XLSX}")
    except PermissionError:
        wb.save(str(XLSX_OUT))
        print(f"\nArquivo salvo em: {XLSX_OUT}")
        print("ATENÇÃO: feche o Excel e copie esse arquivo para substituir o original:")
        print(f"  {XLSX}")
    print("Aba '14_Dest_CNAE' reconstruída com estrutura País → CNAE.")


if __name__ == "__main__":
    main()
