from __future__ import annotations

from pathlib import Path
import shutil

import pandas as pd
from openpyxl import load_workbook

from comex.reporting.template_report import MONTH_NAMES_PT, TemplateWorkbookData

INVALID_XLSX_CHARS = r'[\[\]\:\*\?\/\\]'


def _is_variation_column(name: str) -> bool:
    normalized = str(name).strip().lower()
    return normalized.startswith("yoy_") or ("variação" in normalized) or ("variacao" in normalized)


def _is_number_like_column(name: str) -> bool:
    normalized = str(name).strip().lower()
    if normalized.endswith("(em us$ fob)") or normalized.endswith("(em quilogramas líquidos)"):
        return True

    return normalized in {
        "nr_competencia",
        "nr_ano",
        "nr_mes",
        "qt_estatistica",
        "qt_kilo_liquido",
        "vl_fob",
        "vl_frete",
        "vl_seguro",
        "row_count",
        "latest_row_count",
        "competencias",
        "flow_kinds",
        "distinct_ncm",
        "distinct_countries",
        "total_vl_fob",
        "total_qt_kilo_liquido",
        "total_qt_estatistica",
        "total_vl_frete",
        "total_vl_seguro",
    }


def _pretty_sheet_title(sheet_name: str) -> str:
    title = sheet_name.replace("_", " ").strip()
    return title.title()


def _autowidth(
    worksheet,
    dataframe: pd.DataFrame,
    start_col: int = 0,
    extra_pad: int = 2,
    max_width: int = 35,
) -> None:
    for index, column in enumerate(dataframe.columns):
        values = dataframe[column].head(500).astype(str).tolist()
        max_len = max([len(str(column)), *[len(value) for value in values]])
        worksheet.set_column(start_col + index, start_col + index, min(max_len + extra_pad, max_width))


def _apply_content_base(worksheet, dataframe: pd.DataFrame, workbook, start_col: int = 0) -> None:
    content_format = workbook.add_format({"font_name": "Aptos Narrow", "font_size": 11})
    worksheet.set_column(start_col, start_col + len(dataframe.columns) - 1, None, content_format)


def _apply_number_formats(worksheet, dataframe: pd.DataFrame, workbook, start_col: int = 0) -> None:
    common = {"font_name": "Aptos Narrow", "font_size": 11}
    integer_format = workbook.add_format({**common, "num_format": "#,##0"})
    decimal_format = workbook.add_format({**common, "num_format": "#,##0.00"})
    percent_format = workbook.add_format({**common, "num_format": "0.0%"})

    numeric_columns = [
        column
        for column in dataframe.columns
        if pd.api.types.is_numeric_dtype(dataframe[column]) or _is_number_like_column(column)
    ]

    for column in numeric_columns:
        column_index = dataframe.columns.get_loc(column)
        normalized = str(column).strip().lower()

        if _is_variation_column(column):
            cell_format = percent_format
        elif normalized.startswith("vl_") or normalized.startswith("total_vl_"):
            cell_format = decimal_format
        elif normalized.startswith("qt_") or normalized.startswith("total_qt_"):
            cell_format = decimal_format
        else:
            cell_format = integer_format

        worksheet.set_column(start_col + column_index, start_col + column_index, None, cell_format)


def _style_headers(worksheet, dataframe: pd.DataFrame, workbook, header_row: int, start_col: int = 0) -> None:
    main_header = workbook.add_format(
        {
            "font_name": "Aptos Narrow",
            "font_size": 12,
            "bold": True,
            "font_color": "white",
            "bg_color": "#0E2841",
            "border": 1,
            "text_wrap": True,
            "align": "center",
            "valign": "vcenter",
        }
    )
    variation_header = workbook.add_format(
        {
            "font_name": "Aptos Narrow",
            "font_size": 12,
            "bold": True,
            "font_color": "white",
            "bg_color": "#215C98",
            "border": 1,
            "text_wrap": True,
            "align": "center",
            "valign": "vcenter",
        }
    )

    worksheet.set_row(header_row, 36)
    for index, column in enumerate(dataframe.columns):
        header_format = variation_header if _is_variation_column(column) else main_header
        worksheet.write(header_row, start_col + index, column, header_format)


def _apply_conditional_formats(
    worksheet,
    dataframe: pd.DataFrame,
    workbook,
    data_first_row: int,
    data_last_row: int,
    start_col: int = 0,
) -> None:
    if data_last_row < data_first_row:
        return

    variation_columns = [column for column in dataframe.columns if _is_variation_column(column)]
    if not variation_columns:
        return

    positive_format = workbook.add_format(
        {
            "font_name": "Aptos Narrow",
            "font_size": 11,
            "bg_color": "#C6EFCE",
            "font_color": "#006100",
        }
    )
    negative_format = workbook.add_format(
        {
            "font_name": "Aptos Narrow",
            "font_size": 11,
            "bg_color": "#FFC7CE",
            "font_color": "#9C0006",
        }
    )

    for column in variation_columns:
        column_index = start_col + dataframe.columns.get_loc(column)
        worksheet.conditional_format(
            data_first_row,
            column_index,
            data_last_row,
            column_index,
            {"type": "cell", "criteria": ">", "value": 0, "format": positive_format},
        )
        worksheet.conditional_format(
            data_first_row,
            column_index,
            data_last_row,
            column_index,
            {"type": "cell", "criteria": "<", "value": 0, "format": negative_format},
        )


def _write_title(worksheet, title_text: str, workbook, column_count: int, row: int = 0) -> None:
    title_format = workbook.add_format(
        {
            "bold": True,
            "font_size": 22,
            "font_name": "Aptos Narrow",
            "font_color": "#0E2841",
            "valign": "vcenter",
        }
    )
    last_col = max(column_count - 1, 0)
    worksheet.merge_range(row, 0, row, last_col, title_text, title_format)


def _write_cover_sheet(writer, workbook, sheets: dict[str, pd.DataFrame]) -> None:
    cover = pd.DataFrame(
        {
            "Item": ["Generated at", "Sheets included"],
            "Value": [
                pd.Timestamp.now().strftime("%Y-%m-%d %H:%M"),
                ", ".join(sheets.keys()),
            ],
        }
    )
    cover.to_excel(writer, sheet_name="Cover", index=False)
    worksheet = writer.sheets["Cover"]
    _autowidth(worksheet, cover)
    cover_format = workbook.add_format({"font_name": "Aptos Narrow", "font_size": 11})
    worksheet.set_column(0, cover.shape[1] - 1, None, cover_format)
    _style_headers(worksheet, cover, workbook, header_row=0)
    worksheet.freeze_panes(1, 0)
    worksheet.autofilter(0, 0, cover.shape[0], cover.shape[1] - 1)


def write_excel_report(output_path: str | Path, sheets: dict[str, pd.DataFrame]) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        workbook = writer.book
        _write_cover_sheet(writer, workbook, sheets)

        for sheet_name, dataframe in sheets.items():
            safe_name = sheet_name[:31]
            start_row = 1
            dataframe.to_excel(writer, sheet_name=safe_name, index=False, startrow=start_row)

            worksheet = writer.sheets[safe_name]
            _write_title(worksheet, _pretty_sheet_title(sheet_name), workbook, len(dataframe.columns), row=0)
            _apply_content_base(worksheet, dataframe, workbook)
            _style_headers(worksheet, dataframe, workbook, header_row=start_row)
            _autowidth(worksheet, dataframe)
            _apply_number_formats(worksheet, dataframe, workbook)

            data_first_row = start_row + 1
            data_last_row = start_row + len(dataframe)
            _apply_conditional_formats(
                worksheet,
                dataframe,
                workbook,
                data_first_row=data_first_row,
                data_last_row=data_last_row,
            )

            worksheet.freeze_panes(start_row + 1, 0)
            worksheet.autofilter(start_row, 0, start_row + len(dataframe), len(dataframe.columns) - 1)

    return output


def _clear_row_values(worksheet, row_number: int, max_column: int) -> None:
    for column_number in range(1, max_column + 1):
        worksheet.cell(row=row_number, column=column_number).value = None


def _summary_descriptions(reference_month: int, reference_year: int) -> list[str]:
    month_name = MONTH_NAMES_PT[reference_month].lower()
    previous_year = reference_year - 1
    return [
        f"Aqui, encontram-se as informações sobre as exportações por produto de {month_name} de {reference_year}, com análises em relação a {month_name} de {previous_year}.",
        f"Aqui, encontram-se as informações sobre as exportações por país de {month_name} de {reference_year}, com análises em relação a {month_name} de {previous_year}.",
        f"Aqui, encontram-se as informações sobre as importações por produto de {month_name} de {reference_year}, com análises em relação a {month_name} de {previous_year}.",
        f"Aqui, encontram-se as informações sobre as importações por país de {month_name} de {reference_year}, com análises em relação a {month_name} de {previous_year}.",
        f"Aqui, encontram-se as informações sobre as exportações por produto do acumulado do ano de {reference_year}, com análises em relação ao mesmo período de {previous_year}.",
        f"Aqui, encontram-se as informações sobre as exportações por país do acumulado do ano de {reference_year}, com análises em relação ao mesmo período de {previous_year}.",
        f"Aqui, encontram-se as informações sobre as importações por produto do acumulado do ano de {reference_year}, com análises em relação ao mesmo período de {previous_year}.",
        f"Aqui, encontram-se as informações sobre as importações por país do acumulado do ano de {reference_year}, com análises em relação ao mesmo período de {previous_year}.",
    ]


def write_template_excel_report(
    template_path: str | Path,
    output_path: str | Path,
    workbook_data: TemplateWorkbookData,
) -> Path:
    template = Path(template_path)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template, output)

    workbook = load_workbook(output)
    data_sheet_names = list(workbook_data.sheet_frames.keys())

    summary_sheet = workbook["Sumário"]
    summary_sheet["C4"] = (
        f"Mês de Referência - ({MONTH_NAMES_PT[workbook_data.reference_month]}/"
        f"{workbook_data.reference_year})"
    )
    summary_sheet["D16"] = (
        f"MDIC ({workbook_data.reference_year}) e Observatório FIESC ({workbook_data.reference_year})"
    )

    descriptions = _summary_descriptions(workbook_data.reference_month, workbook_data.reference_year)

    template_data_sheets = [sheet for sheet in workbook.worksheets if sheet.title != "Sumário"]
    if len(template_data_sheets) < len(data_sheet_names):
        raise RuntimeError("Template workbook does not have enough data sheets.")

    for index, sheet_name in enumerate(data_sheet_names):
        worksheet = template_data_sheets[index]
        worksheet.title = sheet_name
        worksheet["A1"] = workbook_data.titles[sheet_name]

        dataframe = workbook_data.sheet_frames[sheet_name].copy()
        worksheet_max_col = max(worksheet.max_column, len(dataframe.columns))

        for column_index, column_name in enumerate(dataframe.columns, start=1):
            worksheet.cell(row=4, column=column_index).value = column_name

        for row_number in range(5, 15):
            _clear_row_values(worksheet, row_number, worksheet_max_col)

        for row_offset in range(10):
            row_number = 5 + row_offset
            if row_offset >= len(dataframe):
                continue
            row_values = dataframe.iloc[row_offset].tolist()
            for column_index, value in enumerate(row_values, start=1):
                worksheet.cell(row=row_number, column=column_index).value = value

        summary_row = 7 + index
        summary_sheet[f"C{summary_row}"] = sheet_name
        summary_sheet[f"D{summary_row}"] = descriptions[index]

    workbook.save(output)
    return output
